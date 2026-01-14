"""
登录页面测试用例生成器
运行此代码将自动生成完整的测试用例Excel文件
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os


def generate_login_test_cases():
    """生成登录页面测试用例数据"""

    # 测试用例数据
    test_cases_data = {
        "功能测试": [
            ["TC-LOGIN-001", "正常登录流程", "1. 输入已注册的用户名/邮箱\n2. 输入正确的密码\n3. 点击'登录'按钮",
             "1. 登录成功\n2. 跳转到首页或指定页面\n3. 显示欢迎消息", "", "P0", "通过/失败", "核心路径"],
            ["TC-LOGIN-002", "登录会话验证", "1. 成功登录后\n2. 检查浏览器Cookie/Session",
             "1. 正确设置用户会话标识\n2. 没有明文存储敏感信息", "", "P0", "未执行", ""],
            ["TC-LOGIN-003", "用户名为空", "1. 用户名为空\n2. 输入有效密码\n3. 点击登录",
             "1. 显示'用户名不能为空'\n2. 光标定位到用户名输入框", "", "P1", "未执行", ""],
            ["TC-LOGIN-004", "密码为空", "1. 输入有效用户名\n2. 密码为空\n3. 点击登录",
             "1. 显示'密码不能为空'\n2. 光标定位到密码输入框", "", "P1", "未执行", ""],
            ["TC-LOGIN-005", "用户名和密码均为空", "1. 用户名和密码都为空\n2. 点击登录",
             "显示'请输入用户名和密码'提示", "", "P1", "未执行", ""],
            ["TC-LOGIN-006", "错误用户名或密码", "1. 输入错误的用户名或密码\n2. 点击登录",
             "1. 显示'用户名或密码错误'\n2. 不提示具体是哪个错误", "", "P1", "未执行", "安全考虑"],
            ["TC-LOGIN-007", "账户被锁定", "1. 输入已被锁定的账户\n2. 输入任意密码\n3. 点击登录",
             "显示'账户已被锁定，请联系管理员'", "", "P1", "未执行", ""],
            ["TC-LOGIN-008", "密码显示格式", "1. 在密码框输入字符",
             "字符显示为'•'或'*'（密文）", "", "P2", "未执行", ""],
            ["TC-LOGIN-009", "显示/隐藏密码", "1. 输入密码\n2. 点击'显示密码'图标",
             "1. 密码明文显示\n2. 再次点击恢复密文", "", "P2", "未执行", ""],
            ["TC-LOGIN-010", "密码特殊字符", "1. 密码包含特殊字符!@#$%^&*()\n2. 点击登录",
             "登录成功", "", "P2", "未执行", ""],
            ["TC-LOGIN-011", "'记住我'功能", "1. 勾选'记住我'\n2. 登录成功\n3. 关闭浏览器重新打开",
             "自动保持登录状态", "", "P1", "未执行", ""],
            ["TC-LOGIN-012", "不记住登录状态", "1. 不勾选'记住我'\n2. 登录成功\n3. 关闭浏览器重新打开",
             "需要重新登录", "", "P1", "未执行", ""],
            ["TC-LOGIN-013", "忘记密码链接", "1. 点击'忘记密码'链接",
             "跳转到密码重置页面", "", "P1", "未执行", ""],
            ["TC-LOGIN-014", "有效账户找回", "1. 输入已注册的邮箱/手机号\n2. 点击'发送验证码'",
             "1. 收到重置邮件/短信\n2. 提示'验证码已发送'", "", "P1", "未执行", ""],
            ["TC-LOGIN-015", "无效账户找回", "1. 输入未注册的邮箱/手机号\n2. 点击'发送验证码'",
             "提示'该账户不存在'", "", "P1", "未执行", ""],
            ["TC-LOGIN-016", "注册链接", "1. 点击'立即注册'链接",
             "跳转到注册页面", "", "P2", "未执行", ""],
            ["TC-LOGIN-017", "新用户注册后登录", "1. 新用户完成注册\n2. 使用新注册的账户登录",
             "登录成功", "", "P1", "未执行", ""],
            ["TC-LOGIN-018", "连续错误密码锁定", "1. 连续输入错误密码5次\n2. 第6次尝试登录",
             "1. 账户被临时锁定\n2. 显示'请15分钟后再试'或要求验证码", "", "P1", "未执行", ""],
            ["TC-LOGIN-019", "锁定后解锁登录", "1. 账户被锁定后等待解锁时间\n2. 使用正确密码登录",
             "登录成功", "", "P1", "未执行", ""],
            ["TC-LOGIN-020", "第三方登录", "1. 点击'微信登录'按钮\n2. 授权确认",
             "1. 跳转到微信授权页面\n2. 授权后返回并登录成功", "", "P1", "未执行", ""],
        ],

        "安全性测试": [
            ["TC-SEC-001", "HTTPS传输", "1. 检查登录页面URL\n2. 使用抓包工具查看登录请求",
             "1. URL以https://开头\n2. 登录数据加密传输", "", "P0", "未执行", ""],
            ["TC-SEC-002", "XSS攻击防护", "1. 用户名输入<script>alert('xss')</script>\n2. 任意密码登录",
             "1. 脚本被过滤或转义\n2. 不执行JS脚本", "", "P0", "未执行", ""],
            ["TC-SEC-003", "SQL注入防护", "1. 用户名输入' OR '1'='1\n2. 任意密码",
             "1. 登录失败\n2. 不暴露数据库错误信息", "", "P0", "未执行", ""],
            ["TC-SEC-004", "URL敏感信息", "1. 登录成功后检查URL",
             "URL中不包含用户名、密码等敏感信息", "", "P1", "未执行", ""],
            ["TC-SEC-005", "会话固定攻击", "1. 获取登录前的session ID\n2. 登录后检查session ID",
             "登录后session ID发生变化", "", "P1", "未执行", ""],
            ["TC-SEC-006", "验证码有效性", "1. 使用过期的验证码登录",
             "提示'验证码无效或已过期'", "", "P1", "未执行", "如系统有验证码"],
            ["TC-SEC-007", "暴力破解防护", "1. 使用脚本快速尝试不同密码",
             "触发频率限制，需要验证码或临时锁定", "", "P1", "未执行", ""],
            ["TC-SEC-008", "密码传输安全", "1. 使用抓包工具捕获登录请求",
             "密码在前端加密或使用HTTPS传输", "", "P0", "未执行", ""],
        ],

        "UI_UX测试": [  # 修改工作表名，避免/字符
            ["TC-UI-001", "页面布局", "1. 打开登录页面",
             "1. 布局符合设计稿\n2. 元素对齐整齐", "", "P2", "未执行", ""],
            ["TC-UI-002", "焦点状态", "1. 点击用户名输入框",
             "输入框有视觉焦点提示（边框高亮等）", "", "P2", "未执行", ""],
            ["TC-UI-003", "错误提示位置", "1. 触发错误（如空密码）",
             "错误提示显示在对应输入框附近", "", "P2", "未执行", ""],
            ["TC-UI-004", "响应式布局", "1. 在手机、平板、PC上分别访问",
             "页面布局适应不同屏幕尺寸", "", "P1", "未执行", ""],
            ["TC-UI-005", "Tab键导航", "1. 按Tab键在页面元素间切换",
             "1. 焦点按合理顺序移动\n2. 可切换到所有可交互元素", "", "P2", "未执行", ""],
            ["TC-UI-006", "Enter键登录", "1. 输入用户名和密码\n2. 按Enter键",
             "触发登录操作", "", "P2", "未执行", ""],
        ],

        "兼容性测试": [
            ["TC-COM-001", "Chrome浏览器", "1. 在Chrome最新版本测试所有功能",
             "所有功能正常工作", "", "P1", "未执行", ""],
            ["TC-COM-002", "Firefox浏览器", "1. 在Firefox最新版本测试所有功能",
             "所有功能正常工作", "", "P1", "未执行", ""],
            ["TC-COM-003", "Safari浏览器", "1. 在Safari测试所有功能",
             "所有功能正常工作", "", "P1", "未执行", ""],
            ["TC-COM-004", "Edge浏览器", "1. 在Edge浏览器测试所有功能",
             "所有功能正常工作", "", "P1", "未执行", ""],
            ["TC-COM-005", "移动端浏览器", "1. 在iOS Safari和Android Chrome测试",
             "1. 触摸操作正常\n2. 键盘弹出不影响布局", "", "P1", "未执行", ""],
        ],

        "性能测试": [
            ["TC-PER-001", "页面加载性能", "1. 打开登录页面并计时",
             "页面在3秒内完全加载", "", "P2", "未执行", ""],
            ["TC-PER-002", "登录响应时间", "1. 输入正确信息点击登录\n2. 计时到跳转完成",
             "登录操作在2秒内完成", "", "P1", "未执行", ""],
            ["TC-PER-003", "防重复提交", "1. 快速双击登录按钮",
             "只发送一次登录请求", "", "P1", "未执行", ""],
            ["TC-PER-004", "高并发登录", "1. 使用性能测试工具模拟100用户同时登录",
             "1. 成功率>99%\n2. 平均响应时间<3秒", "", "P2", "未执行", ""],
        ],

        "辅助功能测试": [
            ["TC-ACC-001", "屏幕阅读器支持", "1. 使用屏幕阅读器访问登录页面",
             "能正确读取所有标签和提示信息", "", "P2", "未执行", ""],
            ["TC-ACC-002", "键盘操作完整", "1. 不使用鼠标，仅用键盘操作",
             "能完成所有登录操作", "", "P2", "未执行", ""],
            ["TC-ACC-003", "颜色对比度", "1. 检查文字与背景的对比度",
             "满足WCAG AA标准（4.5:1）", "", "P2", "未执行", ""],
        ]
    }

    return test_cases_data


def create_excel_file(filename="登录页面测试用例.xlsx"):
    """创建并格式化Excel文件"""

    # 获取测试用例数据
    test_cases_data = generate_login_test_cases()

    # 创建Excel写入器
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        workbook = writer.book

        # 为每个测试类型创建sheet
        for sheet_name, test_cases in test_cases_data.items():
            # 创建DataFrame
            columns = ["用例ID", "测试场景", "测试步骤", "预期结果", "实际结果", "优先级", "执行状态", "备注"]
            df = pd.DataFrame(test_cases, columns=columns)

            # 写入Excel
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # 限制sheet名31字符

            # 获取worksheet进行格式设置
            worksheet = writer.sheets[sheet_name[:31]]

            # 设置列宽
            column_widths = {
                'A': 15,  # 用例ID
                'B': 25,  # 测试场景
                'C': 40,  # 测试步骤
                'D': 40,  # 预期结果
                'E': 15,  # 实际结果
                'F': 10,  # 优先级
                'G': 12,  # 执行状态
                'H': 20  # 备注
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # 设置标题行样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 设置数据行样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 应用样式到所有单元格
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 设置优先级颜色
            for row in range(2, worksheet.max_row + 1):
                priority_cell = worksheet[f'F{row}']
                if priority_cell.value == 'P0':
                    priority_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    priority_cell.font = Font(color="FFFFFF", bold=True)
                elif priority_cell.value == 'P1':
                    priority_cell.fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
                    priority_cell.font = Font(color="FFFFFF", bold=True)
                elif priority_cell.value == 'P2':
                    priority_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 修正这里
                    priority_cell.font = Font(color="000000", bold=True)

        # 创建统计sheet
        create_summary_sheet(workbook, test_cases_data)

    print(f"✅ Excel文件已成功生成: {filename}")
    print(f"📊 文件包含 {len(test_cases_data)} 个分类工作表")
    print(f"📝 总计 {sum(len(cases) for cases in test_cases_data.values())} 个测试用例")

    return filename


def create_summary_sheet(workbook, test_cases_data):
    """创建统计汇总工作表"""

    summary_sheet = workbook.create_sheet(title="测试用例汇总", index=0)

    # 汇总表标题
    summary_sheet['A1'] = "登录页面测试用例汇总报告"
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet.merge_cells('A1:H1')

    # 统计信息
    summary_data = [
        ["分类", "用例数量", "P0", "P1", "P2", "已完成", "未执行", "通过率"],
    ]

    total_cases = 0
    for sheet_name, test_cases in test_cases_data.items():
        count = len(test_cases)
        total_cases += count

        # 统计优先级（简化统计）
        p0_count = sum(1 for case in test_cases if case[5] == 'P0')
        p1_count = sum(1 for case in test_cases if case[5] == 'P1')
        p2_count = sum(1 for case in test_cases if case[5] == 'P2')

        summary_data.append([
            sheet_name, count, p0_count, p1_count, p2_count, 0, count, "0%"
        ])

    # 添加总计行
    summary_data.append([
        "总计", total_cases, "", "", "", 0, total_cases, "0%"
    ])

    # 写入汇总数据
    for row_idx, row_data in enumerate(summary_data, start=3):
        for col_idx, cell_value in enumerate(row_data, start=1):
            summary_sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # 设置汇总表格式
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 12
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        summary_sheet.column_dimensions[col].width = 10

    # 设置标题行样式
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in summary_sheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 设置总计行样式
    total_row = 3 + len(test_cases_data)
    for col in range(1, 9):
        cell = summary_sheet.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        if col == 1:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 添加生成时间
    from datetime import datetime
    generated_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_sheet.cell(row=total_row + 2, column=1, value=f"生成时间: {generated_time}")

    # 添加说明
    summary_sheet.cell(row=total_row + 4, column=1, value="优先级说明:")
    summary_sheet.cell(row=total_row + 5, column=1, value="P0: 核心功能，必须测试")
    summary_sheet.cell(row=total_row + 6, column=1, value="P1: 重要功能，建议测试")
    summary_sheet.cell(row=total_row + 7, column=1, value="P2: 次要功能，可选测试")


def main():
    """主函数"""
    print("🚀 开始生成登录页面测试用例Excel文件...")
    print("=" * 50)

    try:
        # 检查依赖库
        try:
            import pandas as pd
            from openpyxl import Workbook
        except ImportError as e:
            print("❌ 缺少必要的依赖库，请运行以下命令安装：")
            print("   pip install pandas openpyxl")
            return

        # 生成文件名
        filename = "登录页面完整测试用例.xlsx"

        # 创建Excel文件
        created_file = create_excel_file(filename)

        print("=" * 50)
        print("🎉 文件生成成功！")
        print("\n📋 包含的工作表:")
        print("   1. 测试用例汇总 - 统计概览")
        print("   2. 功能测试 - 20个测试用例")
        print("   3. 安全性测试 - 8个测试用例")
        print("   4. UI_UX测试 - 6个测试用例")
        print("   5. 兼容性测试 - 5个测试用例")
        print("   6. 性能测试 - 4个测试用例")
        print("   7. 辅助功能测试 - 3个测试用例")
        print(f"\n💾 文件位置: {os.path.abspath(created_file)}")

        print("\n📝 使用说明:")
        print("   1. 打开Excel文件，在'执行状态'列填写测试结果")
        print("   2. 在'实际结果'列记录测试实际输出")
        print("   3. '汇总'工作表会自动统计测试进度")
        print("   4. 根据项目需要，可以添加'测试人员'、'执行日期'等列")

    except Exception as e:
        print(f"❌ 生成文件时出错: {str(e)}")
        print("\n🔧 常见问题解决:")
        print("   1. 确保已安装依赖库: pip install pandas openpyxl")
        print("   2. 关闭正在使用的Excel文件")
        print("   3. 检查文件写入权限")
        print("   4. 确保磁盘有足够空间")


if __name__ == "__main__":
    main()